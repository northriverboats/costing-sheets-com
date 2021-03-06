import click
import sys
import os
import re
from pickle import load
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color, colors
from dotenv import load_dotenv

dbg = 0
consumable_column = 9
without_options = ""
with_options = " WITH OPTIONS"
yellow = None
yellow_fill = None
status = None
bundle_dir = "."

rates = [
    # labor rate, column, row
    ["FABRICATION", 7, 428],
    ["PAINT", 7, 429],
    ["CANVAS", 7, 430],
    ["OUTFITTING", 7, 431],
]
sections = [
    # Section, start, end, consumable, start-del, end-del, markup
    ["FABRICATION", 8, 32, 35, 0, 0, 0.0],
    ["PAINT", 42, 64, 67, 0, 0, 0.0],
    ["CANVAS", 74, 96, 0, 70, 102, 0.0],
    ["OUTFITTING", 106, 356, 0, 0, 0, 0.0],
    ["ENGINE & JET", 391, 401, 0, 387, 407, 0.0],
    ["TRAILER", 411, 411, 0, 0, 0, 0.0],
]


class Status():
    def __init__(self):
        self.__percent = 0
        self.__file = ''
        self.__section = ''

    def __str__(self):
        return "{}\t{}\t{}".format(
            self.__percent, 
            self.__file,
            self.__section)

    def output(self):
        print(self)

    @property
    def percent(self):
        return self.__percent

    @percent.setter
    def percent(self, value):
        self.__percent = value

    @property
    def file(self):
        return self.__file

    @file.setter
    def file(self, value):
        self.__file = value

    @property
    def section(self):
        return self.__section

    @section.setter
    def section(self, value):
        self.__section = value



"""
Levels
0 = no output
1 = show model/cabin
2 = show lengths
3 = show sections
4 = show part merges
5 = show part adds
"""
def debug(level, text):
    if dbg > (level -1):
        print(text)

def load_environment():
    global bundle_dir
    if getattr(sys, 'frozen', False):
        bundle_dir = sys._MEIPASS # pylint: disable=no-member
    else:
        # we are running in a normal Python environment
        bundle_dir = os.path.dirname(os.path.abspath(__file__))

    # load environmental variables
    load_dotenv(dotenv_path = Path(bundle_dir) / ".env")

def unpickle_boats(pickle_folder):
    base = Path(pickle_folder)

    pickle_file = base.joinpath(base.stem + '.pickle')
    if pickle_file.exists() == False:
        debug(0, str(pickle_file) + " not found.")
        sys.exit(1)

    boats = load(open(pickle_file, 'rb'))
    return boats

def resolve_environment(value, environmental_string):
    if value == None:
        value = os.getenv(environmental_string)
    return value

def setup_styles():
    global yellow, yellow_fill
    yellow = Color(colors.YELLOW)
    yellow_fill = PatternFill(fill_type='solid', start_color=yellow, end_color=yellow)

# Delete extra blank lines in section
def delete_unused_section(ws, start_delete_row, end_delete_row):
    range_to_move = "A{}:I{}".format(
        end_delete_row,
        ws.max_row + 300
    )
    ws.move_range(
        range_to_move,
        rows=start_delete_row - end_delete_row,
        cols=0,
        translate=False
    )
    recalc_sheet(ws,start_delete_row, start_delete_row - end_delete_row)

def process_labor_rate_hours(ws, boats, model, length):
    for rate, column, row in rates:
        labor = float(boats[model][rate + " LABOR RATE"])
        hours = boats[model][rate + " " + str(length) + " HOURS"]
        if hours == "#N/A":
            hours = 0.0
        else:
            hours = float(hours)
        _ = ws.cell(column=column, row=row, value=labor)
        _ = ws.cell(column=column-1, row=row, value=hours)

def process_part_highlighting(ws, length, part, mode, sheet_type, row):
    if sheet_type == without_options:
        return
    if "P" in mode:
        ws.cell(row=row, column=2).fill = yellow_fill
    if "Z" in mode:
        ws.cell(row=row, column=1).fill = yellow_fill
        ws.cell(row=row, column=2).fill = yellow_fill

def process_by_parts(ws, length, parts, start, end, markup):
    offset = 0
    for part in sorted(parts, key = lambda i: (str(i['VENDOR']), str(i['PART NUMBER']))):
        qty = float(part['QTY'])
        row = start + offset
        if qty > 0:
            ws.cell(column=1, row=row, value=part['VENDOR'])
            ws.cell(column=2, row=row, value=part['PART NUMBER'][1:-1])
            if part['DESCRIPTION'] != 'do not use':
                ws.cell(column=3, row=row, value=part['DESCRIPTION'])
            price = float(part['PRICE']) + float(part['PRICE']) * markup
            ws.cell(column=4, row=row, value=price)
            ws.cell(column=5, row=row, value=part['UOM'])
            ws.cell(column=6, row=row, value=qty)
            offset += 1

    delete_unused_section(ws, start + offset, end)
 
def adjust_formula(function, threshold, offset):
    pattern =  re.compile(r"(\$?[A-Za-z]{1,3})(\$?[1-9][0-9]{0,6})")
    start_pos = 0
    result = ""
    for match in pattern.finditer(function):
        result += function[start_pos:match.start()+1]
        num = match.group(2)
        if int(num) >= threshold:
            num = str(int(num) + offset)
        result += num
        start_pos = match.end()
    if start_pos <= len(function):
        result += function[start_pos:]
    return result

def recalc_sheet(ws, threshold, offset):
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if str(value)[0] == "=":
                formula = adjust_formula(value, threshold, offset)
                if formula != value:
                    cell.value = formula

def delete_unused_materials_and_labor_rate(ws, section):
    if section == 'ENGINE & JET':
        return None
    cells = [cell for cell in ws['D'] if cell.value == section.title()]
    for cell in cells:
        row = cell.row
        delete_unused_section(ws, row, row + 1)

def process_consumables(ws, boats, model, length, section, start_row, consumable_row):
    if consumable_row > 0:
        consumables = float(boats[model][section + ' CONSUMABLES'])
        formula = "=SUM({}{}:{}{})*{}".format(
            chr(64 + consumable_column),
            start_row - 1,
            chr(64 + consumable_column),
            consumable_row - 1,
            consumables
        )
        _ = ws.cell(column=consumable_column, row=consumable_row, value=formula)

def  add_to_section_parts(parts, part, msg):
    # implictly passing reference to parts list
    if float(part['QTY']) == 0.0:
        return
    # add to current item if it exists
    found_items = [x for x in parts if x['PART NUMBER'] == part['PART NUMBER']]
    if found_items:
        debug(4, '        Now Merging {}: {}'.format(msg, part['PART NUMBER']))
        found_items[0]['QTY'] = float(found_items[0]['QTY']) + float(part['QTY'])
    else:
        debug(5, '        Now adding {}: {}'.format(msg, part['PART NUMBER']))
        parts.append(part)

def parts_list_for_section(parts, boats, model, length, section, msg):
    # implictly passing reference to parts list
    for master_part in boats[model][section + ' PARTS']:
        part = {x: y for x, y in master_part.items() if x in [
            'PART NUMBER',
            'DESCRIPTION',
            'UOM',
            'PRICE',
            'VENDOR',
            'VENDOR PART'
        ]}
        part['QTY'] = master_part[str(length) + ' QTY']
        part['TOTAL'] = master_part[str(length) + ' TOTAL']
        add_to_section_parts(parts, part, msg)

def process_by_section(ws, base, boats, model_name, model, length, sheet_type):
    for section, start_row, end_row, consumable_row, start_delete_row, end_delete_row, markup in sections[::-1]:
        debug(3, '      {}'.format(section))
        parts = []  # implictly passing reference to parts list
        parts_list_for_section(parts, boats, model, length, section, "part")
        if not base:
            parts_list_for_section(parts, boats, model_name + ' MAIN', length, section, "main")
        number_of_parts = len(parts)
        if number_of_parts == 0 and start_delete_row > 0:
            delete_unused_section(ws, start_delete_row, end_delete_row)
            delete_unused_materials_and_labor_rate(ws, section)
        else:
            process_consumables(ws, boats, model, length, section, start_row, consumable_row)
            process_by_parts(ws, length, parts, start_row, end_row, markup)

def generate_filename(folder, base, model, length, sheet_type):
    tag = sheet_type
    # if base:
    #   tag += " with Hours"
    return folder + "\\Parts Costing {}' {} 2020{}.xlsx".format(length, model.upper(), tag)

def create_folder_if_needed(folder):
    if not os.path.exists(folder):
        os.makedirs(folder)

def save_spreadsheet(wb, base, model, length, sheet_type, output_folder):
    folder = output_folder + "\\" + model.upper()
    create_folder_if_needed(folder)
    file_name = generate_filename(folder, base, model, length, sheet_type)
    wb.save(file_name)

def process_sheetname(ws, model, length):
    value = "{}' {}".format(length, model)
    debug(2,"    {}".format(length))
    _ = ws.cell(column=3, row=3, value=value)

def load_template(template_file):
    wb = load_workbook(Path(bundle_dir) / template_file)
    ws = wb.active
    return [wb, ws]

def set_print_range(ws):
    max_row = [cell for cell in ws['G'] if cell.value =="CONTRIBUTION MARGIN"][0].row
    print_area = "A1:I{}".format(max_row)
    ws.print_area = print_area

def process_boat(base, boats, model_name, model, length, output_folder, template_file, sheet_type):
    wb, ws = load_template(template_file)
    
    process_sheetname(ws, model, length)
    process_labor_rate_hours(ws, boats, model, length)
    process_by_section(ws, base, boats, model_name, model, length, sheet_type)
    set_print_range(ws)

    save_spreadsheet(wb, base, model, length, sheet_type, output_folder)

def process_by_length(base, boats, model_name, model, output_folder, template_file):
    for length in boats[model]["BOAT SIZES"]:
        process_boat(base, boats, model_name, model, length, output_folder, template_file, without_options)

def process_by_cabin(base, boats, model_name, output_folder, template_file):
    for model in [x for x in boats if model_name in x]:
        if not base and "MAIN" in model:
            continue
        debug(1, '  {}'.format(model))
        process_by_length(base, boats, model_name, model, output_folder, template_file)

def process_by_model(base, boats, output_folder, template_file):
    # get model name by finding MAIN
    for model_name in [x[:-5] for x in boats if 'MAIN' in x]:
        debug(1, '{}'.format(model_name))
        process_by_cabin(base, boats, model_name, output_folder, template_file)

def setup_debug(verbose, machine):
    global dbg, status
    status = Status()
    if not machine:
        dbg = verbose


def setup_markups(fabrication, paint, canvas, outfitting, engine, trailer):
    global sections
    for section_name, markup in [
        ['FABRICATION', fabrication],
        ['PAINT', paint],
        ['CANVAS', canvas],
        ['OUTFITTING', outfitting],
        ['ENGINE & JET', engine],
        ['TRAILER', trailer],
    ]:
        [section for section in sections if section[0] == section_name][0][6] = markup


# pylint: disable=no-value-for-parameter
@click.command()
@click.option('--verbose', '-v', default=0, type=int, help='verbosity level 0-3')
@click.option('--machine', '-m', default=0, type=int, help='machine readable output, overrides --verbose')
@click.option('--base', '-b', is_flag=True, help='merge base model MAIN with each sections parts')
@click.option('--markup-fabrication', '-mf', default=0, type=click.FloatRange(-1, 1), help='markup 0.05 is 5%')
@click.option('--markup-paint', '-mp', default=0, type=click.FloatRange(-1, 1), help='markup 0.05 is 5%')
@click.option('--markup-canvas', '-mc', default=0, type=click.FloatRange(-1, 1), help='markup 0.05 is 5%')
@click.option('--markup-outfitting', '-mo', default=0, type=click.FloatRange(-1, 1), help='markup 0.05 is 5%')
@click.option('--markup-engine', '-me', default=0, type=click.FloatRange(-1, 1), help='markup 0.05 is 5%')
@click.option('--markup-trailer', '-mt', default=0, type=click.FloatRange(-1, 1), help='markup 0.05 is 5%')
@click.option('--folder', '-f', required=False, type=click.Path(exists=True, file_okay=False), help="directory to process")
@click.option('--output', '-o', required=False, type=click.Path(exists=True, file_okay=False), help="output directory")
@click.option('--template', '-t', required=False, type=click.Path(exists=True, dir_okay=False), help="template xlsx sheet")
def main(
    verbose,
    machine,
    base,
    markup_fabrication,
    markup_paint,
    markup_canvas,
    markup_outfitting,
    markup_engine,
    markup_trailer,
    folder,
    output,
    template
):
    setup_debug(verbose, machine)
    setup_markups(markup_fabrication, markup_paint, markup_canvas, markup_outfitting, markup_engine, markup_trailer)
    load_environment()
    pickle_folder = resolve_environment(folder, 'FOLDER')
    output_folder = resolve_environment(output, 'OUTPUT')
    template_file = resolve_environment(template, 'TEMPLATE')

    boats = unpickle_boats(pickle_folder)
    setup_styles()
    process_by_model(base, boats, output_folder, template_file)


if __name__ == "__main__":
    main()
